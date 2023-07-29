from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


name_file = "/home/user/PycharmProjects/WEB_parsing/others/exzemple.xlsx"
wb = load_workbook(name_file)
ws = wb.active
for row in range(2,30):
    for col in range(2, 5):
        char = get_column_letter(col)
        print(ws[char + str(row)].value)




# wb = Workbook()
# ws = wb.active
# ws.title = 'Data'
#
# ws.append(['Tim', 'Is'])
#
# wb.save('/home/user/PycharmProjects/WEB_parsing/others/tim.xlsx')





# name_file = "/home/user/PycharmProjects/WEB_parsing/others/exzemple.xlsx"
# wb = load_workbook(name_file)
# ws = wb.active

# ws["A1"].value = "test" #   присваивает значение ячейке
# wb.create_sheet("test")     #   создаст новый лист
# wb.save(name_file)







